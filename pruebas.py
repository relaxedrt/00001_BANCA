#Importamos las librerias necesarias
import openpyxl
import secrets

#Creamos la ruta a la base de datos
databasedir = "database.xlsx"

#Creamos la variable general del usuario activo y de administrador
activeuser = ""
admin = False

#Con esta función crearemos un usuario nuevo siempre y cuando no exista
#uno bajo el mismo dni, en dicho caso devolveremos el error por pantalla y 
#volveremos a la instancia principal.

def register():
    #Abrimos la base de datos
    databasexlsx = openpyxl.load_workbook(databasedir)
    database = databasexlsx["login"]
    print("====================================================")
    print("//                   ARTIBANK SL                  //")
    print("====================================================")

    #Pedimos el dni al usuario
    user = input("-Introduzca su dni: ")

    #Comparamos el dni con la base de datos
    for dni in range(2, 1048576):
        #Si ya existe en nuestra base de datos activamos el bit de existencia
        if database.cell(row=dni, column=1).value == user:
            print("El dni introducido ya tiene una cuenta: ")
            dniexists = True
            break
        else:
            dniexists = False
    #Si no existia el dni en nuestra base de datos pedimos la contraseña
    if not dniexists:
        pwd1 = input("-Introduzca su contraseña: ")
        pwd2 = input("-Vuelva a introducir su contraseña: ")
        if pwd1 == pwd2:
            pwdmissmatch = False
            for dni in range(2, 1048576):
                if database.cell(row=dni, column=1).value == None:
                    database.cell(row=dni, column=1).value = user
                    database.cell(row=dni, column=2).value = pwd2
                    databasexlsx.save(databasedir) 
                    break
        else:
            print("Las contraseñas no coinciden.")
            pwdmissmatch = True
            
    if dniexists or pwdmissmatch:
        print("Ha habido un error registrando el usuario.")

#Con esta función nos loguearemos con un usuario siempre y cuando exista
#uno bajo ese dni, en dicho caso devolveremos el error por pantalla y 
#volveremos a la instancia principal.
        
def login():
    #Accedemos a las variables globales
    global activeuser
    global admin

    #Abrimos la base de datos
    databasexlsx = openpyxl.load_workbook(databasedir)
    database = databasexlsx["login"]
    print("====================================================")
    print("//                   ARTIBANK SL                  //")
    print("====================================================")

    #Pedimos el dni y la contraseña al usuario
    user = input("-Introduzca su dni: ")
    pwd =  input("-Introduzca su contraseña: ")

    #Comparamos los valores con los de la base de datos
    for dni in range(2, 1048576):
        if database.cell(row=dni, column=1).value == user:
            if database.cell(row=dni, column=2).value == pwd:
                #Si la contraseña hace match con ese user devolvemos success
                activeuser = user
                if database.cell(row=dni, column=3).value == 1:
                    admin = True
                print("success")
            else:
                #Si la contraseña no hace match con ese user devolvemos password para identificar el error
                print("wrong password")
    #Si recorremos toda la base de datos y no encontramos el dni devolvemos user para saber que no existe
    print("user doesn't exists")

#Con esta función nos desloguearemos de nuestra sesión actual

def logout():
    #Accedemos a las variables globales
    global activeuser
    global admin

    activeuser = ""
    admin = False

    print("logged out")

#Con esta función crearemos una cuenta a nombre del usuario activo
#o bajo el del usuario deseado con el admin

def createaccount():
    #Accedemos a las variables globales
    global activeuser
    global admin

    #Abrimos la base de datos
    databasexlsx = openpyxl.load_workbook(databasedir)
    database = databasexlsx["accounts"]
    users = databasexlsx["login"]

    print("====================================================")
    print("//                   ARTIBANK SL                  //")
    print("====================================================")
    
    if admin:
        user = input("A que usuario desea crearle una cuenta?: ")
        for usr in range(2, 1048576):
            #Comprobamos que exista ese usuario
            if users.cell(row=usr, column=1).value == user:
                for acc in range(2, 1048576):
                    if database.cell(row=acc, column=1).value == None:
                        database.cell(row=acc, column=1).value = secrets.token_hex(16)
                        database.cell(row=acc, column=2).value = int("0")
                        database.cell(row=acc, column=3).value = user
                        print(f"La nueva cuenta de {user} es: {database.cell(row=acc, column=1).value}")
                        databasexlsx.save(databasedir) 
                        return 0
    else:
        for acc in range(2, 1048576):
            if database.cell(row=acc, column=1).value == None:
                database.cell(row=acc, column=1).value = secrets.token_hex(16)
                database.cell(row=acc, column=2).value = int("0")
                database.cell(row=acc, column=3).value = activeuser
                print(f"Tu nueva cuenta es: {database.cell(row=acc, column=1).value}")
                databasexlsx.save(databasedir) 
                return 0
    print("No existe ese usuario.")

def checkaccounts():
    #Accedemos a las variables globales
    global activeuser
    global admin

    #Abrimos la base de datos
    databasexlsx = openpyxl.load_workbook(databasedir)
    database = databasexlsx["accounts"]
    users = databasexlsx["login"]

    print("====================================================")
    print("//                   ARTIBANK SL                  //")
    print("====================================================")
    
    if admin:
        user = input("A que usuario desea ver las cuentas?: ")
        for usr in range(2, 1048576):
            #Comprobamos que exista ese usuario
            if users.cell(row=usr, column=1).value == user:
                #Recorremos la lista de cuentas
                for acc in range(2, 1048576):
                    #Si el usuario coincide con el dueño de la cuenta mostramos los datos por pantalla
                    if database.cell(row=acc, column = 3) == user:
                        print(f"La cuenta {database.cell(row=acc, column = 1).value} del usuario {user} tiene {database.cell(row=acc, column = 2)}")
    else:
        #Recorremos la lista de cuentas
        for acc in range(2, 1048576):
            #Si el usuario coincide con el dueño de la cuenta mostramos los datos por pantalla
            if database.cell(row=acc, column = 3) == activeuser:
                print(f"La cuenta {database.cell(row=acc, column = 1).value} del usuario {activeuser} tiene {database.cell(row=acc, column = 2)}")
    print("No existe ese usuario.")

def deposit():
    #Accedemos a las variables globales
    global activeuser
    global admin

    #Abrimos la base de datos
    databasexlsx = openpyxl.load_workbook(databasedir)
    database = databasexlsx["accounts"]
    users = databasexlsx["login"]

    print("====================================================")
    print("//                   ARTIBANK SL                  //")
    print("====================================================")
    
    if admin:
        user = input("¿En qué usuario quiere realizar el depósito?: ")
        for usr in range(2, 1048576):
            #Comprobamos que exista ese usuario
            if users.cell(row=usr, column=1).value == user:
                #Recorremos la lista de cuentas
                for acc in range(2, 1048576):
                    #Si el usuario coincide con el dueño de la cuenta mostramos los datos por pantalla
                    if database.cell(row=acc, column = 3) == user:
                        print(f"La cuenta {database.cell(row=acc, column = 1).value} del usuario {user} tiene {database.cell(row=acc, column = 2)}")
        
        #Preguntamos a que cuenta ingresar y cuanto ingresar
        account = input("¿En qué cuenta quieres ingresar el dinero?: ")
        money = int(input("¿Cuanto dinero quires ingresar?: "))

        #Buscamos la cuenta indicada
        for acc in range(2, 1048576):
            if database.cell(row=acc, column = 1) == account:
                #Actualizamos la columna del dinero
                database.cell(row=acc, column = 2).value += money
                databasexlsx.save(databasedir) 
                return 0
    else:
        #Recorremos la lista de cuentas
        for acc in range(2, 1048576):
            #Si el usuario coincide con el dueño de la cuenta mostramos los datos por pantalla
            if database.cell(row=acc, column = 3) == activeuser:
                print(f"La cuenta {database.cell(row=acc, column = 1).value} del usuario {activeuser} tiene {database.cell(row=acc, column = 2)}")
            
        #Preguntamos a que cuenta ingresar y cuanto ingresar
        account = input("¿En qué cuenta quieres ingresar el dinero?: ")
        money = int(input("¿Cuanto dinero quires ingresar?: "))

        #Buscamos la cuenta indicada
        for acc in range(2, 1048576):
            if (database.cell(row=acc, column = 1).value == account) and (database.cell(row=acc, column = 3) == activeuser):
                #Actualizamos la columna del dinero
                database.cell(row=acc, column = 2).value += money
                databasexlsx.save(databasedir)
                return 0
            elif (database.cell(row=acc, column = 1).value == account) and (database.cell(row=acc, column = 3) != activeuser):
                print("Esa no es tu cuenta. Intenta nuevamente.")
                return 0
    print("No existe ese usuario.")

def withdrawal():
    #Accedemos a las variables globales
    global activeuser
    global admin

    #Abrimos la base de datos
    databasexlsx = openpyxl.load_workbook(databasedir)
    database = databasexlsx["accounts"]
    users = databasexlsx["login"]

    print("====================================================")
    print("//                   ARTIBANK SL                  //")
    print("====================================================")
    
    if admin:
        user = input("¿En qué usuario quiere realizar el depósito?: ")
        for usr in range(2, 1048576):
            #Comprobamos que exista ese usuario
            if users.cell(row=usr, column=1).value == user:
                #Recorremos la lista de cuentas
                for acc in range(2, 1048576):
                    #Si el usuario coincide con el dueño de la cuenta mostramos los datos por pantalla
                    if database.cell(row=acc, column = 3) == user:
                        print(f"La cuenta {database.cell(row=acc, column = 1).value} del usuario {user} tiene {database.cell(row=acc, column = 2)}")
        
        #Preguntamos a que cuenta retirar y cuanto retirar
        account = input("¿De qué cuenta quieres retirar el dinero?: ")
        money = int(input("¿Cuanto dinero quires retirar?: "))

        #Buscamos la cuenta indicada
        for acc in range(2, 1048576):
            if database.cell(row=acc, column = 1) == account:
                if database.cell(row=acc, column = 2).value >= money:
                    #Actualizamos la columna del dinero
                    database.cell(row=acc, column = 2).value -= money
                    databasexlsx.save(databasedir)
                    return 0
                else:
                    print("No hay suficiente dinero.")
                    return 0
    else:
        #Recorremos la lista de cuentas
        for acc in range(2, 1048576):
            #Si el usuario coincide con el dueño de la cuenta mostramos los datos por pantalla
            if database.cell(row=acc, column = 3) == activeuser:
                print(f"La cuenta {database.cell(row=acc, column = 1).value} del usuario {activeuser} tiene {database.cell(row=acc, column = 2)}")
            
        #Preguntamos a que cuenta retirar y cuanto retirar
        account = input("¿De qué cuenta quieres retirar el dinero?: ")
        money = int(input("¿Cuanto dinero quires retirar?: "))

        #Buscamos la cuenta indicada
        for acc in range(2, 1048576):
            if (database.cell(row=acc, column = 1).value == account) and (database.cell(row=acc, column = 3) == activeuser):
                if database.cell(row=acc, column = 2).value >= money:
                    #Actualizamos la columna del dinero
                    database.cell(row=acc, column = 2).value -= money
                    databasexlsx.save(databasedir)
                    return 0
                else:
                    print("No hay suficiente dinero.")
                    return 0
            elif (database.cell(row=acc, column = 1).value == account) and (database.cell(row=acc, column = 3) != activeuser):
                print("Esa no es tu cuenta. Intenta nuevamente.")
                return 0
    print("No existe ese usuario.")


#Ciclo de programa
while True:
    #Si no hay un usuario activo, se muestra el menú de login/register
    if  activeuser == "":  
        print("====================================================")
        print("//                   ARTIBANK SL                  //")
        print("//  1.Iniciar sesión                              //")
        print("//  2.Registrar usuario                           //")
        print("//  3.Cerrar aplicación                           //")
        print("====================================================")

        opt = input("¿Que opción deseas ejecutar?: ")

        if  int(opt)==1:
            login()
        elif int(opt)==2:
            register()
        elif int(opt)==3:
            exit()
        else:
            print("No es una opción válida.\n")
    #Si hay un usuario activo, se muestra el menú de login/register
    else:
        print("====================================================")
        print("//                   ARTIBANK SL                  //")
        print(f"//                   {activeuser}                  //")
        print("//                                                //")
        print("//  1.Chequear cuentas        4.Depositar         //")
        print("//  2.Crear cuentas           5.Retirar           //")
        print("//  3.Saldo total             6.Transferencia     //")
        print("====================================================")
        print("//  98.Logout                 99.EXIT             //")
        print("====================================================")

        opt = input("¿Que opción deseas ejecutar?: ")

        if  int(opt)==1:
            checkaccounts()
        elif int(opt)==2:
            createaccount()
        #elif int(opt)==3:
            #total()
        elif int(opt)==4:
            deposit()
        elif int(opt)==5:
            withdrawal()
        #elif int(opt)==6:
            #transfer()
        elif int(opt)==98:
            logout()
        elif int(opt)==99:
            exit()
        else:
            print("No es una opción válida.\n")